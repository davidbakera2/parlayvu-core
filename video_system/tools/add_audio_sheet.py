from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


HEADERS = [
    "enabled",
    "audio_id",
    "anchor",
    "file",
    "start",
    "end",
    "duration",
    "source_start",
    "volume",
    "fade_in",
    "fade_out",
    "duck_under_voice",
    "notes",
]


def seconds_to_time(seconds: float) -> str:
    total_ms = round(seconds * 1000)
    h, rem = divmod(total_ms, 3600000)
    m, rem = divmod(rem, 60000)
    s, ms = divmod(rem, 1000)
    return f"{h:02d}:{m:02d}:{s:02d}.{ms:03d}"


def time_to_seconds(value: str) -> float:
    text = str(value or "").replace(",", ".")
    if not text:
        return 0.0
    hms, *ms_part = text.split(".")
    h, m, s = [int(part) for part in hms.split(":")]
    ms = int((ms_part[0] if ms_part else "0").ljust(3, "0")[:3])
    return h * 3600 + m * 60 + s + ms / 1000


def load_scenes(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Scenes" not in wb.sheetnames:
        return []
    ws = wb["Scenes"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or "").strip().lower().replace(" ", "_") for cell in rows[0]]
    scenes = []
    for row in rows[1:]:
        record = {header: value for header, value in zip(headers, row) if header}
        if any(value not in ("", None) for value in record.values()):
            scenes.append(record)
    return scenes


def scene_end(scene: dict[str, str]) -> float:
    end = scene.get("end")
    if end:
        return time_to_seconds(str(end))
    start = time_to_seconds(str(scene.get("start") or ""))
    duration = time_to_seconds(str(scene.get("duration") or ""))
    return start + duration


def default_rows(path: Path, music_duration: float) -> list[list[object]]:
    scenes = load_scenes(path)
    if not scenes:
        project_end = 0.0
    else:
        project_end = max(scene_end(scene) for scene in scenes)

    intro_end = 0.0
    for scene in scenes:
        layout = str(scene.get("layout") or "").strip().lower()
        if layout in {"intro", "show_image"}:
            intro_end = max(intro_end, scene_end(scene))
        else:
            break
    if intro_end <= 0:
        intro_end = min(27.0, project_end or 27.0)

    outro_duration = min(10.0, project_end) if project_end else 10.0
    outro_start = max(0.0, project_end - outro_duration)
    outro_source_start = max(0.0, music_duration - outro_duration)

    return [
        [
            True,
            "intro_music",
            "intro_start",
            "music.wav",
            "00:00:00.000",
            "00:00:50.000",
            "00:00:50.000",
            "00:00:00.000",
            0.34,
            "00:00:00.000",
            "00:00:03.000",
            True,
            "Music under intro and opening show image.",
        ],
        [
            True,
            "outro_music",
            "outro_start",
            "music.wav",
            "-00:00:15.000",
            "00:00:05.000",
            "00:00:20.000",
            seconds_to_time(max(0.0, music_duration - 20.0)),
            0.72,
            "00:00:03.000",
            "00:00:01.250",
            False,
            "Music returns for outro.",
        ],
    ]


def add_audio_sheet(workbook_path: Path, music_duration: float) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    if "Audio" in wb.sheetnames:
        del wb["Audio"]
    insert_at = wb.sheetnames.index("Broll") + 1 if "Broll" in wb.sheetnames else len(wb.sheetnames)
    ws = wb.create_sheet("Audio", insert_at)

    ws.append(HEADERS)
    for row in default_rows(workbook_path, music_duration):
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    widths = {
        "A": 10,
        "B": 18,
        "C": 22,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 16,
        "I": 10,
        "J": 14,
        "K": 14,
        "L": 18,
        "M": 38,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{ws.max_row}"

    bool_validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(bool_validation)
    bool_validation.add("A2:A200")
    bool_validation.add("L2:L200")

    wb.save(workbook_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Add an Audio sheet to a video planning workbook.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--music-duration", type=float, default=124.91)
    args = parser.parse_args()
    add_audio_sheet(args.workbook, args.music_duration)
    print(args.workbook)


if __name__ == "__main__":
    main()
