from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, time, timedelta
from pathlib import Path


TIME_RE = re.compile(r"^(?P<h>\d{1,2}):(?P<m>\d{2}):(?P<s>\d{2})(?:[.,](?P<ms>\d{1,3}))?$")


def load_workbook(path: Path):
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl is required for spreadsheet conversion in this local workflow.") from exc
    return openpyxl.load_workbook(path, data_only=True)


def as_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_bool(value) -> bool:
    if value is None or value == "":
        return True
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() not in {"false", "0", "no", "n", "disabled"}


def normalize_time(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}:{value.second:02d}.{int(value.microsecond / 1000):03d}"
    if isinstance(value, timedelta):
        total_ms = int(value.total_seconds() * 1000)
        h, rem = divmod(total_ms, 3600000)
        m, rem = divmod(rem, 60000)
        s, ms = divmod(rem, 1000)
        return f"{h:02d}:{m:02d}:{s:02d}.{ms:03d}"
    if isinstance(value, (int, float)):
        total_ms = int(round(value * 24 * 3600 * 1000))
        h, rem = divmod(total_ms, 3600000)
        m, rem = divmod(rem, 60000)
        s, ms = divmod(rem, 1000)
        return f"{h:02d}:{m:02d}:{s:02d}.{ms:03d}"
    text = str(value).strip().replace(",", ".")
    sign = ""
    if text.startswith("-"):
        sign = "-"
        text = text[1:].strip()
    match = TIME_RE.match(text)
    if match:
        ms = (match.group("ms") or "0").ljust(3, "0")[:3]
        return f"{sign}{int(match.group('h')):02d}:{int(match.group('m')):02d}:{int(match.group('s')):02d}.{ms}"
    return f"{sign}{text}"


def time_to_ms(value: str) -> int | None:
    if not value:
        return None
    match = TIME_RE.match(value)
    if not match:
        return None
    ms = int((match.group("ms") or "0").ljust(3, "0")[:3])
    return ((int(match.group("h")) * 60 + int(match.group("m"))) * 60 + int(match.group("s"))) * 1000 + ms


def ms_to_time(value: int) -> str:
    h, rem = divmod(max(0, value), 3600000)
    m, rem = divmod(rem, 60000)
    s, ms = divmod(rem, 1000)
    return f"{h:02d}:{m:02d}:{s:02d}.{ms:03d}"


def sheet_rows(wb, sheet_name: str) -> list[dict[str, object]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [as_text(cell).lower().replace(" ", "_") for cell in rows[0]]
    output: list[dict[str, object]] = []
    for row in rows[1:]:
        if all(value is None or as_text(value) == "" for value in row):
            continue
        record = {}
        for header, value in zip(headers, row):
            if not header:
                continue
            if header in {
                "start",
                "end",
                "duration",
                "broll_source_start",
                "default_source_start",
                "source_start",
                "fade_in",
                "fade_out",
            }:
                record[header] = normalize_time(value)
            elif header == "enabled":
                record[header] = as_bool(value)
            else:
                record[header] = as_text(value)
        if any(value not in {"", None} for value in record.values()):
            if sheet_name == "Scenes" and not record.get("duration"):
                start_ms = time_to_ms(str(record.get("start", "")))
                end_ms = time_to_ms(str(record.get("end", "")))
                if start_ms is not None and end_ms is not None:
                    record["duration"] = ms_to_time(end_ms - start_ms)
            output.append(record)
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert video_plan.xlsx to video_plan.json.")
    parser.add_argument("project", help="Path to project folder")
    args = parser.parse_args()

    project = Path(args.project).resolve()
    workbook = project / "planning" / "video_plan.xlsx"
    wb = load_workbook(workbook)

    plan = {
        "project": project.name,
        "scenes": sheet_rows(wb, "Scenes"),
        "graphics": sheet_rows(wb, "Graphics"),
        "broll": sheet_rows(wb, "Broll"),
        "audio": sheet_rows(wb, "Audio"),
        "assets": sheet_rows(wb, "Assets"),
        "settings": sheet_rows(wb, "Settings"),
    }

    out = project / "planning" / "video_plan.json"
    out.write_text(json.dumps(plan, indent=2), encoding="utf-8")
    print(out)


if __name__ == "__main__":
    main()
