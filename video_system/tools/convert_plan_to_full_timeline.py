from __future__ import annotations

import argparse
import json
import shutil
import subprocess
from datetime import datetime, time, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


PROGRAM_LAYOUTS = {"1cam", "2cam", "2cam_broll", "3cam", "3cam_broll"}
BOOKEND_LAYOUTS = {"intro", "show_image", "outro"}


def find_tool(name: str) -> str:
    found = shutil.which(name)
    return found or name


FFPROBE = find_tool("ffprobe")


def parse_time(value, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1_000_000
    if isinstance(value, timedelta):
        return value.total_seconds()
    if isinstance(value, (int, float)):
        return float(value) * 24 * 3600
    text = str(value).strip().replace(",", ".")
    if not text or text.startswith("="):
        return default
    parts = text.split(".")
    hms = parts[0].split(":")
    if len(hms) != 3:
        return default
    h, m, s = (int(part) for part in hms)
    ms = int((parts[1] if len(parts) > 1 else "0").ljust(3, "0")[:3])
    return h * 3600 + m * 60 + s + ms / 1000


def fmt(seconds: float) -> str:
    total_ms = round(max(0.0, seconds) * 1000)
    h, rem = divmod(total_ms, 3_600_000)
    m, rem = divmod(rem, 60_000)
    s, ms = divmod(rem, 1000)
    return f"{h:02d}:{m:02d}:{s:02d}.{ms:03d}"


def media_duration(path: Path) -> float:
    if not path.exists():
        return 0.0
    proc = subprocess.run(
        [FFPROBE, "-v", "error", "-show_entries", "format=duration", "-of", "json", str(path)],
        check=True,
        capture_output=True,
        text=True,
    )
    return float(json.loads(proc.stdout)["format"]["duration"])


def headers_for(ws) -> list[str]:
    return [str(cell.value or "").strip().lower().replace(" ", "_") for cell in ws[1]]


def row_to_record(ws, row_index: int, headers: list[str]) -> dict[str, object]:
    return {header: ws.cell(row_index, col_index).value for col_index, header in enumerate(headers, start=1) if header}


def record_is_blank(record: dict[str, object]) -> bool:
    return all(value is None or str(value).strip() == "" for value in record.values())


def set_sheet_rows(ws, headers: list[str], records: list[dict[str, object]]) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row_index, record in enumerate(records, start=2):
        for col_index, header in enumerate(headers, start=1):
            ws.cell(row_index, col_index).value = record.get(header)
    for table in ws.tables.values():
        table.ref = f"A1:{get_column_letter(ws.max_column)}{max(1, len(records) + 1)}"
        if table.autoFilter:
            table.autoFilter.ref = table.ref


def setting_map(wb) -> dict[str, object]:
    ws = wb["Settings"]
    rows = {}
    for row in ws.iter_rows(min_row=2):
        key = row[0].value
        if key:
            rows[str(key).strip()] = row[1].value
    return rows


def upsert_setting(wb, key: str, value: str, note: str) -> None:
    ws = wb["Settings"]
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value or "").strip() == key:
            row[1].value = value
            if ws.max_column >= 3:
                row[2].value = note
            return
    ws.append([key, value, note])


def convert_workbook(project: Path) -> None:
    workbook = project / "planning" / "video_plan.xlsx"
    assets = project / "assets"
    wb = openpyxl.load_workbook(workbook)
    settings = setting_map(wb)
    if str(settings.get("timeline_mode") or "").strip().lower() == "full_rendered":
        print(f"{workbook} already uses full rendered-video timeline mode.")
        return

    scenes_ws = wb["Scenes"]
    scene_headers = headers_for(scenes_ws)
    scene_records = [
        row_to_record(scenes_ws, row_index, scene_headers)
        for row_index in range(2, scenes_ws.max_row + 1)
    ]
    scene_records = [record for record in scene_records if not record_is_blank(record)]
    layouts = {str(record.get("layout") or "").strip() for record in scene_records}

    if not (layouts & BOOKEND_LAYOUTS):
        intro_asset = str(settings.get("intro_asset") or "intro.mp4").strip()
        show_asset = str(settings.get("show_image_asset") or "show_image.png").strip()
        intro_duration = media_duration(assets / intro_asset)
        opening_duration = parse_time(settings.get("opening_show_image_duration"), 5.0)
        outro_duration = parse_time(settings.get("outro_show_image_duration"), 5.0)

        converted: list[dict[str, object]] = []
        converted.append(
            {
                "enabled": True,
                "scene_id": "INTRO",
                "start": fmt(0),
                "end": fmt(intro_duration),
                "duration": fmt(intro_duration),
                "layout": "intro",
                "host_source": intro_asset,
                "notes": "Automatic opening intro, now visible in full rendered timeline.",
            }
        )
        converted.append(
            {
                "enabled": True,
                "scene_id": "OPENING_SHOW_IMAGE",
                "start": fmt(intro_duration),
                "end": fmt(offset),
                "duration": fmt(opening_duration),
                "layout": "show_image",
                "host_source": show_asset,
                "notes": "Opening show image, now visible in full rendered timeline.",
            }
        )

        cursor = intro_duration + opening_duration
        for record in scene_records:
            if str(record.get("layout") or "").strip() not in PROGRAM_LAYOUTS:
                continue
            original_start = parse_time(record.get("start"))
            duration = max(0.0, parse_time(record.get("end"), original_start) - original_start)
            if record.get("source_start") in {None, ""}:
                record["source_start"] = fmt(original_start)
            start = cursor
            end = cursor + duration
            record["start"] = fmt(start)
            record["end"] = fmt(end)
            record["duration"] = fmt(duration)
            cursor = end
            converted.append(record)

        converted.append(
            {
                "enabled": True,
                "scene_id": "OUTRO",
                "start": fmt(cursor),
                "end": fmt(cursor + outro_duration),
                "duration": fmt(outro_duration),
                "layout": "outro",
                "host_source": show_asset,
                "notes": "Automatic outro show image, now visible in full rendered timeline.",
            }
        )
        set_sheet_rows(scenes_ws, scene_headers, converted)

        if "Graphics" in wb.sheetnames:
            graphics_ws = wb["Graphics"]
            graphic_headers = headers_for(graphics_ws)
            graphic_records = [
                row_to_record(graphics_ws, row_index, graphic_headers)
                for row_index in range(2, graphics_ws.max_row + 1)
            ]
            graphic_records = [record for record in graphic_records if not record_is_blank(record)]
            offset = intro_duration + opening_duration
            for record in graphic_records:
                if record.get("start") not in {None, ""}:
                    record["start"] = fmt(parse_time(record.get("start")) + offset)
                if record.get("end") not in {None, ""}:
                    record["end"] = fmt(parse_time(record.get("end")) + offset)
            set_sheet_rows(graphics_ws, graphic_headers, graphic_records)
    else:
        first_program = next((record for record in scene_records if str(record.get("layout") or "").strip() in PROGRAM_LAYOUTS), None)
        interview_start = parse_time(first_program.get("start")) if first_program else 0.0
        cursor = 0.0
        normalized: list[dict[str, object]] = []
        for record in scene_records:
            layout = str(record.get("layout") or "").strip()
            if layout not in (PROGRAM_LAYOUTS | BOOKEND_LAYOUTS):
                continue
            duration = parse_time(record.get("duration"))
            if duration <= 0:
                duration = max(0.0, parse_time(record.get("end")) - parse_time(record.get("start")))
            if layout in PROGRAM_LAYOUTS and record.get("source_start") in {None, ""}:
                record["source_start"] = fmt(max(0.0, parse_time(record.get("start")) - interview_start))
            record["start"] = fmt(cursor)
            record["end"] = fmt(cursor + duration)
            record["duration"] = fmt(duration)
            cursor += duration
            normalized.append(record)
        set_sheet_rows(scenes_ws, scene_headers, normalized)

        if "Graphics" in wb.sheetnames and first_program:
            program_records = [
                record
                for record in normalized
                if str(record.get("layout") or "").strip() in PROGRAM_LAYOUTS
            ]

            def source_to_final(source_time: float) -> float:
                for program in program_records:
                    program_source = parse_time(program.get("source_start"))
                    program_start = parse_time(program.get("start"))
                    program_duration = parse_time(program.get("duration"))
                    if program_source <= source_time <= program_source + program_duration:
                        return program_start + (source_time - program_source)
                return source_time + interview_start

            graphics_ws = wb["Graphics"]
            graphic_headers = headers_for(graphics_ws)
            graphic_records = [
                row_to_record(graphics_ws, row_index, graphic_headers)
                for row_index in range(2, graphics_ws.max_row + 1)
            ]
            graphic_records = [record for record in graphic_records if not record_is_blank(record)]
            for record in graphic_records:
                if record.get("start") not in {None, ""}:
                    source_start = max(0.0, parse_time(record.get("start")) - interview_start)
                    record["start"] = fmt(source_to_final(source_start))
                if record.get("end") not in {None, ""}:
                    source_end = max(0.0, parse_time(record.get("end")) - interview_start)
                    record["end"] = fmt(source_to_final(source_end))
            set_sheet_rows(graphics_ws, graphic_headers, graphic_records)

    upsert_setting(wb, "auto_intro", "FALSE", "Bookends are now explicit rows in Scenes.")
    upsert_setting(wb, "auto_opening_show_image", "FALSE", "Bookends are now explicit rows in Scenes.")
    upsert_setting(wb, "auto_outro_show_image", "FALSE", "Bookends are now explicit rows in Scenes.")
    upsert_setting(wb, "timeline_mode", "full_rendered", "Scenes.start/end match the final rendered-video timeline.")

    wb.save(workbook)
    print(workbook)


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert a video_plan.xlsx to full rendered-video timeline mode.")
    parser.add_argument("project", help="Path to project folder")
    args = parser.parse_args()
    convert_workbook(Path(args.project).resolve())


if __name__ == "__main__":
    main()
